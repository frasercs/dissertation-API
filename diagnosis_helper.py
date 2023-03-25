"""
A helper file used to perform the calculations and get the data for the diagnosis_controller.py file.
"""

import os
import sys

from openpyxl import load_workbook
from werkzeug.exceptions import BadRequest

# load the Excel file
wb = load_workbook(filename=os.path.join(sys.path[0], "data.xlsx"))


@staticmethod
def validate_priors(priors, diseases):
    """
    A function used to validate the priors provided by the user
    :param priors: A dictionary of priors for each disease, where the key is the disease and the value is the prior
    :param diseases: A list of the diseases that are valid for the animal
    :return: The priors dictionary if it is valid, otherwise a BadRequest exception is raised
    """
    provided_keys = []
    for key in priors.keys():
        if key not in diseases:
            raise BadRequest(f"Disease '{key}' is not a valid disease. Please use a valid disease from {diseases}.")
        provided_keys.append(key)

    for disease in diseases:
        if disease not in provided_keys:
            raise BadRequest(
                f"Missing '{disease}' in priors. Please provide a prior likelihood value for all diseases.")

    total_value = sum(priors.values())
    if total_value != 100:
        raise BadRequest(f"Priors must add up to 100. Currently they add up to {total_value}.")

    return priors


@staticmethod
def validate_likelihoods(likelihoods, diseases, signs):
    """
    A function used to validate the likelihoods provided by the user
    :param likelihoods: A dictionary of likelihoods for each disease, where the key is the disease and the value is a
    dictionary of likelihoods for each sign, where the key is the sign and the value is the likelihood
    :param diseases: A list of the diseases that are valid for the animal
    :param signs: A list of the signs that are valid for the animal
    :return: A dictionary of likelihoods for each disease, where the key is the disease and the value is a
    """
    provided_keys = []
    likelihoods = likelihoods
    for key in likelihoods.keys():
        if key not in diseases:
            raise BadRequest(f"Disease '{key}' in \'likelihoods\' is not a valid disease.")
        provided_keys.append(key)

    for disease in diseases:
        if disease not in provided_keys:
            raise BadRequest(f"Missing '{disease}' in likelihoods. Please provide a likelihood value for all diseases.")

    for disease in likelihoods:
        current_likelihoods = likelihoods[disease]
        provided_keys = []
        for key in current_likelihoods.keys():
            if key not in signs:
                raise BadRequest(f"Sign '{key}' in {disease} within \'likelihoods\' is not a valid sign. "
                                 f"Please use a valid signs from {signs}.")
            provided_keys.append(key)

        for sign in signs:
            if sign not in provided_keys:
                raise BadRequest(
                    f"Missing '{sign}' in likelihoods for disease '{disease}'. Please provide a likelihood value "
                    f"for all signs.")

        for sign in current_likelihoods:
            if (not current_likelihoods[sign] > 0) or (current_likelihoods[sign] >= 1):
                raise BadRequest(f"Likelihood for sign '{sign}' in disease '{disease}' is not a valid value. "
                                 f"Please use a value greater than 0 and less than 1.")

    return likelihoods


@staticmethod
def calculate_results(diseases, likelihoods, shown_signs, priors):
    """
    A function used to calculate the results of the Bayes Theorem
    :param diseases: A list of the diseases that are valid for the animal
    :param likelihoods: A dictionary of likelihoods for each disease, where the key is the disease and the value is a
    dictionary of likelihoods for each sign, where the key is the sign and the value is the likelihood
    :param shown_signs: A dictionary of signs that are shown, where the key is the sign and the value is the presence
    :param priors: A dictionary of priors for each disease, where the key is the disease and the value is the prior
    :return: A dictionary of results for each disease, where the key is the disease and the value is the result
    """
    results = {}
    for disease in diseases:
        chain_probability = 1.0
        current_likelihoods = likelihoods[disease]
        for s in shown_signs:
            presence = shown_signs[s]
            if presence == 1:
                chain_probability *= current_likelihoods[s]
            elif presence == -1:
                chain_probability *= (1 - current_likelihoods[s])
        posterior = chain_probability * priors[disease]
        results[disease] = posterior * 100
    return results


@staticmethod
def normalise(results):
    """
    A function used to normalise the results of the Bayes Theorem calculations
    :param results: A dictionary of results for each disease, where the key is the disease and the value is the result
    :return: A dictionary of normalised results for each disease, where the key is the disease and the value is the
    normalised result
    """
    normalised_results = {}
    summed_results = sum(results.values())
    for r in results:
        value = results[r]
        norm = value / summed_results
        normalised_results[r] = norm * 100
    return normalised_results


@staticmethod
def get_default_priors(diseases):
    """
    A function used to generate equal priors if the user does not provide any
    :param diseases: A list of the diseases that are valid for the animal
    :return priors: A dictionary of priors for each disease, where the key is the disease and the value is the prior
    """
    priors = {}
    for disease in diseases:
        priors[disease] = 100 / len(diseases)
    return priors


@staticmethod
def get_disease_wiki_ids(animal):
    """
    A function used to get the WikiData IDs for the diseases
    :param animal: The animal that is being diagnosed
    :return: A dictionary of WikiData IDs for each disease, where the key is the disease and the value is the
    WikiData ID
    """
    # Empty dictionary to store the WikiData IDs
    wiki_ids = {}
    # Get the list of diseases
    for disease in get_diseases(animal):
        # Loop through the diseases in the Excel sheet
        for row in wb['Disease_Codes'].rows:
            # If the disease is found, add it to the dictionary
            if row[0].value == disease:
                wiki_ids[disease] = row[1].value
    # Return the dictionary
    return wiki_ids


# A function used to collect the data from the Excel workbook which I manually formatted to ensure the data works.
def get_likelihood_data(animal):
    """
    A function used to get the likelihood data from the Excel workbook
    :param animal: The animal that is being diagnosed
    :return: A dictionary of likelihoods for each disease, where the key is the disease and the value is a
    dictionary of likelihoods for each sign, where the key is the sign and the value is the likelihood
    """
    # Load the correct Excel sheet
    ws = wb[animal]
    # Get the list of signs and diseases
    signs = get_signs(animal)
    diseases = get_diseases(animal)
    # Dictionary which stores the likelihoods of each disease
    likelihoods = {}
    # Loop through all rows in the workbook
    for i, row in enumerate(ws.rows):
        # Skip the first row as it is just the headers
        if i == 0:
            continue
        # Dictionary which stores the likelihoods of each sign for the current disease
        current_disease_likelihoods = {}
        # Loop through all cells in each row except the first one as it is the disease name
        for j, cell in enumerate(row[1:]):
            chance = cell.value
            current_disease_likelihoods[signs[j]] = chance
        likelihoods[diseases[i - 1]] = current_disease_likelihoods
    return likelihoods


@staticmethod
def get_diseases(animal):
    """
    A function used to get the diseases from the Excel workbook
    :param animal: The animal that is being diagnosed
    :return: A list of diseases
    """
    # List which stores every given disease
    diseases = []
    ws = wb[animal]
    # Populate the list of diseases
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        for cell in row:
            diseases.append(cell.value)
    return diseases


@staticmethod
def get_signs(animal):
    """
    A function used to get the signs from the Excel workbook
    :param animal: The animal that is being diagnosed
    :return: A list of signs
    """
    # List which stores every given sign
    signs = []
    ws_signs = wb[animal]
    # Populate the list of signs
    for col in ws_signs.iter_cols(min_col=2, max_row=1):
        for cell in col:
            signs.append(cell.value)
    return signs


@staticmethod
def get_sign_names_and_codes(animal):
    """
    Get the full name and Wikidata code for each sign
    :param animal: The animal that is being diagnosed
    :return: A dictionary of full sign data, where the key is the sign and the value is a dictionary of the full name
    and Wikidata code
    """
    ws = wb[animal + '_Abbr']
    full_sign_data = {}
    medical_name = {}
    wikidata_code = {}
    for row in ws.rows:
        medical_name["name"] = row[1].value
        wikidata_code["code"] = row[2].value
        full_sign_data[row[0].value] = {**medical_name, **wikidata_code}
    return full_sign_data


def validate_animal(animal):
    """
    This function is used to validate the animal parameter in the endpoints, it is a global function as it is used in
    multiple endpoints and in multiple functions in both the getHelper and diseaseHelper classes.
    :param animal: The animal that is being diagnosed
    :return bool/str: False if the animal is not valid, the animal name if it is valid
    """
    valid_animals = [name for name in wb.sheetnames if "_Abbr" not in name and "_Codes" not in name]
    animal = animal.capitalize()
    if animal not in valid_animals:
        return False
    else:
        return animal
