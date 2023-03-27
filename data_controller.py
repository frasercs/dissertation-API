import os
import random
import sys

from flask import jsonify
from flask_restx import Namespace, Resource
from openpyxl import load_workbook


import diagnosis_helper as dh

api = Namespace('data', description='Data related operations')

# load the Excel file
wb = load_workbook(filename=os.path.join(sys.path[0], "data.xlsx"))


@api.route('/full_animal_data/<string:animal>')
@api.doc(example='Goat', required=True,
         responses={200: 'OK', 400: 'Bad Request', 404: 'Not Found', 500: 'Internal Server Error'},
         params={'animal': 'The species of animal you wish to retrieve signs and diseases for. This must be a valid '
                           'animal as returned by /data/valid_animals. \n \n'},
         description='<h1>Description</h1><p>This endpoint returns a '
                     '<a href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> '
                     'object containing diagnosable diseases with their corresponding '
                     '<a href="https://www.wikidata.org/">WikiData IDs</a> as well as the valid signs associated with '
                     'the animal, their full medical terminology in English, and their corresponding '
                     '<a href="https://www.wikidata.org/">WikiData IDs</a> (if they exist).</p>'
                     '<h1>URL Parameters</h1><ul><li><p>animal: The species of animal you wish to retrieve signs and '
                     'diseases for. This must be a valid animal as returned by /data/valid_animals. '
                     '</p></li></ul>\n \n ')
class getRequiredInputData(Resource):
    """
    This class is used to create the full_animal_data endpoint which returns the required input data for the
    diagnose endpoint.
    """

    @staticmethod
    def get(animal):
        # This is the GET method for the full_animal_data endpoint which returns the required input data for the
        # diagnose endpoint.

        animal = dh.validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /data/valid_animals.', 'status': 404}, 404

        return jsonify(
            {'diseases': dh.get_disease_wiki_ids(animal), 'signs': dh.get_sign_names_and_codes(animal)})


@api.hide
@api.route('/matrix/<string:animal>')
@api.doc(example='Goat', required=True,
         responses={200: 'OK', 400: 'Bad Request', 404: 'Not Found', 500: 'Internal Server Error'}, params={
        'animal': 'The species of animal you wish to retrieve the disease sign matrix for. This must be a valid '
                  'animal as returned by /data/valid_animals. \n \n'},
         description='<h1>Description</h1><p>This endpoint returns a '
                     '<a href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> '
                     'object containing the disease-sign Bayesian matrix for the given animal. This matrix contains '
                     'the likelihoods of each sign being present for each disease.</p><h1>URL Parameters</h1><ul><li>'
                     '<p>animal: The species of animal you wish to retrieve the disease-sign matrix for. '
                     'This must be a valid animal as returned by /data/valid_animals. </p></li></ul>\n \n ')
class getDiseaseSignMatrix(Resource):
    """
    This class is used to create the matrix endpoint which returns the disease sign matrix for the given animal. It is
    hidden from the Swagger UI as it is not intended to be used by the user, it is only to be used by the developer
    or specific users permitted in order to update the data on mobile apps which need to perform calculations offline.
    """


    @staticmethod
    def get(animal):
        # This is the GET method for the matrix endpoint which returns the disease sign matrix for the given animal.
        # It is hidden from the Swagger UI as it is not intended to be used by the user, it is only to be used by the
        # developer or specific users permitted in order to update the data on mobile apps which need to perform
        # calculations offline.

        animal = dh.validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /data/valid_animals.', 'status': 404}, 404

        return jsonify(dh.get_likelihood_data(animal))


@api.route('/example_matrix/<string:animal>')
@api.doc(example='Sheep', required=True,
         responses={200: 'OK', 400: 'Bad Request', 404: 'Not Found', 500: 'Internal Server Error'}, params={
        'animal': 'The species of animal you wish to retrieve the disease sign matrix for. This must be a valid '
                  'animal as returned by /data/valid_animals. \n \n'},
         description='<h1>Description</h1><p>This endpoint returns a '
                     '<a href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> '
                     'object containing an example disease-sign Bayesian matrix for the given animal. '
                     'This matrix contains randomly generated '
                     'likelihoods of each sign being present for each disease.</p><h1>URL Parameters</h1><ul><li>'
                     '<p>animal: The species of animal you wish to retrieve the disease-sign matrix for. '
                     'This must be a valid animal as returned by /data/valid_animals. </p></li></ul>\n \n ')
class getExampleMatrix(Resource):
    """
    This class is used to create the example_matrix endpoint which returns a randomly generated disease sign matrix
    for the given animal. It is used for documentation and giving context to users of the API as the data used in
    the diagnosis process is not available to the public.
    """


    @staticmethod
    def get(animal):
        # This is the GET method for the example_matrix endpoint

        animal = dh.validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /data/valid_animals.', 'status': 404}, 404

        data = dh.get_likelihood_data(animal)

        # Replace correct data with random data
        for disease in data:
            for sign in data[disease]:
                data[disease][sign] = round(random.uniform(0.00001, 0.99999), 4)

        return jsonify(data)


@api.route('/valid_animals')
@api.doc(responses={200: 'OK', 500: 'Internal Server Error'}, description='<h1>Description</h1>'
                                                                          '<p>This endpoint returns a <a '
                                                                          'href="https://developer.mozilla.org/en-US/docs'
                                                                          '/Learn/JavaScript'
                                                                          '/Objects/JSON">JSON</a> object containing the '
                                                                          'names of animal species which are available for '
                                                                          'diagnosis in the /api/diagnose POST method '
                                                                          'below. </p>\n')
class getAnimals(Resource):
    """
    This class is used to create the valid_animals endpoint which returns the names of animals which are available for
    diagnosis in the /diagnosis/diagnose POST method
    """

    @staticmethod
    def get():
        # This is the GET method for the valid_animals endpoint

        # Return the names by getting every worksheet with doesn't contain _Abbr or _Codes
        # I'm using list comprehension as it is cleaner than a for loop, but syntacitically slightly more advanced

        return jsonify([name for name in wb.sheetnames if "_Abbr" not in name and "_Codes" not in name])


@api.route('/full_sign_data/<string:animal>')
@api.doc(required=True, responses={200: 'OK', 400: 'Bad Request', 404: 'Not Found', 500: 'Internal Server Error'},
         description='<h1>Description</h1>'
                     '<p>This endpoint returns a <a '
                     'href="https://developer.mozilla'
                     '.org/en-US/docs/Learn/JavaScript'
                     '/Objects/JSON">JSON</a> object '
                     'which contains the full medical '
                     'terminology for each sign'
                     ' in English as '
                     'well as the corresponding <a '
                     'href="https://www.wikidata.org'
                     '/">WikiData IDs</a> for the '
                     'signs ('
                     'if they exist).</p>'
                     '<h1>URL Parameters</h1>'
                     '<ul>'
                     '<li><p>animal: The species of '
                     'animal you wish to retrieve '
                     'signs and diseases for. This '
                     'must be '
                     'a valid '
                     'animal as returned by '
                     '/data/valid_animals.</p></li>'
                     '</ul>',
         params={'animal': 'The species of animal you wish to retrieve the data for. This must be a valid animal as '
                           'returned by /data/valid_animals. \n \n '})
class getSignCodesAndTerminology(Resource):
    """
    This class is used to create the full_sign_data endpoint which returns the full medical terminology for each sign
    in English as well as the corresponding WikiData IDs for the signs (if they exist).
    """


    @staticmethod
    def get(animal):
        # This is the GET method for the full_sign_data endpoint

        animal = dh.validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /data/valid_animals.', 'status': 404}, 404

        return jsonify({'full_sign_data': dh.get_sign_names_and_codes(animal)})


@api.route('/full_disease_data/<string:animal>')
@api.doc(required=True, responses={200: 'OK', 400: 'Bad Request', 404: 'Not Found', 500: 'Internal Server Error'},
         description='<h1>Description</h1>'
                     '<p>This endpoint returns a <a '
                     'href="https://developer.mozilla'
                     '.org/en-US/docs/Learn/JavaScript'
                     '/Objects/JSON">JSON</a> object '
                     'which contains the possible '
                     'diseases for the given animal as '
                     'well as the corresponding <a '
                     'href="https://www.wikidata.org'
                     '/">WikiData IDs</a> (if they '
                     'exist).</p>'
                     '<h1>URL Parameters</h1>'
                     '<ul>'
                     '<li><p>animal: The species of '
                     'animal you wish to retrieve '
                     'signs and diseases for. This '
                     'must be '
                     'a valid '
                     'animal as returned by '
                     '/data/valid_animals.</p></li>'
                     '</ul>',
         params={'animal': 'The species of animal you wish to retrieve the data for. This must be a valid animal as '
                           'returned by'
                           '/data/valid_animals. \n \n'})
class getDiseaseCodes(Resource):
    """
    This class is used to create the full_disease_data endpoint which returns the possible diseases for the given
    animal as well as the corresponding WikiData IDs (if they exist).
    """

    @staticmethod
    def get(animal):
        # This is the GET method for the full_disease_data endpoint

        animal = dh.validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /data/valid_animals.', 'status': 404}, 404

        return jsonify({'disease_codes': dh.get_disease_wiki_ids(animal)})