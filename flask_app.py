"""

This file contains the code for the Flask API. It is used to diagnose animals using Bayes Theorem.

"""

import os
import random
import sys
from typing import Dict, List, Union

from flask import Flask, request, jsonify, Response
from flask_cors import CORS
from flask_restx import Api, Resource, fields
from openpyxl import load_workbook
from werkzeug.exceptions import BadRequest

# load the Excel file
wb = load_workbook(filename=os.path.join(sys.path[0], "data.xlsx"))

# init the api
api = Api(version='1.0', title='Diagnosis API', description='A simple API to diagnose animals using Bayes Theorem.',
          default='Diagnosis API', default_label='Diagnosis API')
# init the flask app
app = Flask(__name__)
# fix the cors issue
CORS(app)
# init the api using factory pattern
api.init_app(app)

diagnosis_payload_model = api.model('Diagnose', {
    'animal': fields.String(required=True, description='The species of animal. As of version 1.0 this can be'
                                                       ' \'Cattle\', \'Sheep\',\'Goat\', \'Camel\', \'Horse\'or '
                                                       '\'Donkey\'.', example='Cattle'),
    'signs': fields.Raw(required=True,
                        description='The signs shown by the animal. For example: {\"Anae\": 0, \"Anrx\":1,'
                                    '\"Atax\": 0, \"Const\": 0,\"Diarr\": 0, \"Dysnt\": 1, \"Dyspn\": 0,'
                                    '\"Icter\": 0, \"Lymph\": -1,\"Pyrx\": 0, \"Stare\": 0, \"Stunt\": 0,'
                                    '\"SV_Oedm\": 1, \"Weak\": 0,\"Wght_L\": 0}',
                        example={"Anae": 0, "Anrx": 1, "Atax": 0, "Const": 0, "Diarr": 0, "Dysnt": 1, "Dyspn": 0,
                                 "Icter": 0, "Lymph": -1, "Pyrx": 0, "Stare": 0, "Stunt": 0, "SV_Oedm": 1, "Weak": 0,
                                 "Wght_L": 0}), 'priors': fields.Raw(required=False,
                                                                     description='This field can be used if you wish '
                                                                                 'to specify which diseases are more '
                                                                                 'likely than others. If left blank, '
                                                                                 'the algorithm will assume all '
                                                                                 'diseases have an equal chance of '
                                                                                 'occurring. The values given MUST '
                                                                                 'add up to 100. The values given are '
                                                                                 'the percentage likelihood of each '
                                                                                 'disease occurring. \n \n '
                                                                                 'You must provide data for every '
                                                                                 'disease. This is the case as '
                                                                                 'without every disease being '
                                                                                 'considered, the algorithm\'s output '
                                                                                 'will be far less accurate',
                                                                     example={"Anthrax": 5, "Babesiosis": 5,
                                                                              "Blackleg": 5, "CBPP": 5,
                                                                              "Colibacillosis": 5, "Cowdriosis": 5,
                                                                              "FMD": 5, "Fasciolosis": 5, "LSD": 5,
                                                                              "Lungworm": 25, "Pasteurollosis": 5,
                                                                              "PGE / GIT parasite": 5, "Rabies": 5,
                                                                              "Trypanosomosis": 5, "Tuberculosis": 5,
                                                                              "ZZ_Other": 5}),
    'likelihoods': fields.Raw(required=False,
                              description='This field can be used to define your own likelihood data for each disease '
                                          'being the cause of each sign.If left blank, the algorithm will use the '
                                          'default likelihood data. The values given MUST add up to be greater than '
                                          '0 and less than 1. The values given are the percentage likelihood of each '
                                          'sign being present when the animal has the disease. \n \n You must provide '
                                          'data for every disease and every sign. This is the case as without '
                                          'every disease and sign being considered, the algorithm\'s '
                                          'will not work correctly.', example={
            "Anthrax": {"Anae": 0.2668, "Anrx": 0.4023, "Atax": 0.0085, "Const": 0.185, "Diarr": 0.7954,
                "Dysnt": 0.0888, "Dyspn": 0.7864, "Icter": 0.2016, "Lymph": 0.6453, "Pyrx": 0.8447, "SV_Oedm": 0.0814,
                "Stare": 0.5038, "Stunt": 0.7275, "Weak": 0.0771, "Wght_L": 0.6383},
            "Babesiosis": {"Anae": 0.9315, "Anrx": 0.1084, "Atax": 0.5898, "Const": 0.7412, "Diarr": 0.2778,
                "Dysnt": 0.5351, "Dyspn": 0.2371, "Icter": 0.1604, "Lymph": 0.5144, "Pyrx": 0.8542, "SV_Oedm": 0.8295,
                "Stare": 0.8088, "Stunt": 0.9521, "Weak": 0.1691, "Wght_L": 0.6912},
            "Blackleg": {"Anae": 0.2589, "Anrx": 0.1933, "Atax": 0.503, "Const": 0.3529, "Diarr": 0.5158,
                "Dysnt": 0.9027, "Dyspn": 0.3662, "Icter": 0.6998, "Lymph": 0.3139, "Pyrx": 0.2509, "SV_Oedm": 0.735,
                "Stare": 0.1036, "Stunt": 0.5706, "Weak": 0.2188, "Wght_L": 0.0016},
            "CBPP": {"Anae": 0.4047, "Anrx": 0.9005, "Atax": 0.7564, "Const": 0.9049, "Diarr": 0.0014, "Dysnt": 0.0739,
                "Dyspn": 0.9572, "Icter": 0.2247, "Lymph": 0.0228, "Pyrx": 0.8588, "SV_Oedm": 0.2807, "Stare": 0.6536,
                "Stunt": 0.1723, "Weak": 0.0436, "Wght_L": 0.7293},
            "Colibacillosis": {"Anae": 0.5151, "Anrx": 0.6292, "Atax": 0.1373, "Const": 0.4801, "Diarr": 0.5341,
                "Dysnt": 0.8744, "Dyspn": 0.5039, "Icter": 0.7543, "Lymph": 0.4948, "Pyrx": 0.404, "SV_Oedm": 0.5243,
                "Stare": 0.1538, "Stunt": 0.6301, "Weak": 0.8463, "Wght_L": 0.1069},
            "Cowdriosis": {"Anae": 0.1476, "Anrx": 0.7902, "Atax": 0.7319, "Const": 0.7912, "Diarr": 0.7395,
                "Dysnt": 0.6757, "Dyspn": 0.1096, "Icter": 0.3056, "Lymph": 0.185, "Pyrx": 0.3019, "SV_Oedm": 0.4824,
                "Stare": 0.0284, "Stunt": 0.5792, "Weak": 0.1685, "Wght_L": 0.2448},
            "FMD": {"Anae": 0.2977, "Anrx": 0.871, "Atax": 0.963, "Const": 0.1151, "Diarr": 0.9518, "Dysnt": 0.6154,
                "Dyspn": 0.6984, "Icter": 0.4119, "Lymph": 0.3372, "Pyrx": 0.6846, "SV_Oedm": 0.0385, "Stare": 0.6973,
                "Stunt": 0.4824, "Weak": 0.6523, "Wght_L": 0.614},
            "Fasciolosis": {"Anae": 0.0095, "Anrx": 0.977, "Atax": 0.1164, "Const": 0.1311, "Diarr": 0.0569,
                "Dysnt": 0.1604, "Dyspn": 0.0823, "Icter": 0.4349, "Lymph": 0.6208, "Pyrx": 0.9651, "SV_Oedm": 0.1108,
                "Stare": 0.8521, "Stunt": 0.5637, "Weak": 0.1865, "Wght_L": 0.7242},
            "LSD": {"Anae": 0.0883, "Anrx": 0.3197, "Atax": 0.9291, "Const": 0.6622, "Diarr": 0.631, "Dysnt": 0.6657,
                "Dyspn": 0.974, "Icter": 0.153, "Lymph": 0.8023, "Pyrx": 0.2831, "SV_Oedm": 0.4065, "Stare": 0.2959,
                "Stunt": 0.3347, "Weak": 0.5577, "Wght_L": 0.9966},
            "Lungworm": {"Anae": 0.4765, "Anrx": 0.7152, "Atax": 0.6797, "Const": 0.938, "Diarr": 0.6535,
                "Dysnt": 0.9752, "Dyspn": 0.6245, "Icter": 0.4954, "Lymph": 0.4811, "Pyrx": 0.3523, "SV_Oedm": 0.2446,
                "Stare": 0.6479, "Stunt": 0.7919, "Weak": 0.7959, "Wght_L": 0.0079},
            "PGE / GIT parasite": {"Anae": 0.2108, "Anrx": 0.078, "Atax": 0.6913, "Const": 0.664, "Diarr": 0.5129,
                "Dysnt": 0.1369, "Dyspn": 0.27, "Icter": 0.9672, "Lymph": 0.8923, "Pyrx": 0.1874, "SV_Oedm": 0.7678,
                "Stare": 0.4827, "Stunt": 0.6827, "Weak": 0.8251, "Wght_L": 0.4358},
            "Pasteurollosis": {"Anae": 0.5699, "Anrx": 0.8898, "Atax": 0.7183, "Const": 0.9549, "Diarr": 0.4316,
                "Dysnt": 0.5618, "Dyspn": 0.4728, "Icter": 0.5825, "Lymph": 0.2789, "Pyrx": 0.6803, "SV_Oedm": 0.1556,
                "Stare": 0.1606, "Stunt": 0.5968, "Weak": 0.4604, "Wght_L": 0.1261},
            "Rabies": {"Anae": 0.021, "Anrx": 0.1521, "Atax": 0.99, "Const": 0.553, "Diarr": 0.6056, "Dysnt": 0.123,
                "Dyspn": 0.5579, "Icter": 0.8676, "Lymph": 0.4572, "Pyrx": 0.4503, "SV_Oedm": 0.8823, "Stare": 0.6489,
                "Stunt": 0.0886, "Weak": 0.5222, "Wght_L": 0.4999},
            "Trypanosomosis": {"Anae": 0.3602, "Anrx": 0.6961, "Atax": 0.898, "Const": 0.6043, "Diarr": 0.5259,
                "Dysnt": 0.3489, "Dyspn": 0.6187, "Icter": 0.3156, "Lymph": 0.0677, "Pyrx": 0.0405, "SV_Oedm": 0.0257,
                "Stare": 0.5779, "Stunt": 0.9064, "Weak": 0.0779, "Wght_L": 0.9971},
            "Tuberculosis": {"Anae": 0.2207, "Anrx": 0.7513, "Atax": 0.9402, "Const": 0.5418, "Diarr": 0.9419,
                "Dysnt": 0.0402, "Dyspn": 0.6023, "Icter": 0.399, "Lymph": 0.6675, "Pyrx": 0.5587, "SV_Oedm": 0.4947,
                "Stare": 0.59, "Stunt": 0.1479, "Weak": 0.8633, "Wght_L": 0.085},
            "ZZ_Other": {"Anae": 0.6361, "Anrx": 0.1729, "Atax": 0.3008, "Const": 0.6395, "Diarr": 0.3737,
                "Dysnt": 0.858, "Dyspn": 0.9196, "Icter": 0.3255, "Lymph": 0.4681, "Pyrx": 0.7769, "SV_Oedm": 0.0234,
                "Stare": 0.9337, "Stunt": 0.3675, "Weak": 0.1286, "Wght_L": 0.0943}})})

custom_diagnosis_payload_model = api.model('Custom Diagnosis Payload', {

    'diseases': fields.List(fields.String, required=True, description='The diseases to be diagnosed',
                            example=['Rabies', 'Cold']),
    'signs': fields.List(fields.String, required=True, description='The symptoms to be diagnosed',
                         example=['Fever', 'Cough', 'Diarrhoea']),
    'shown_signs': fields.Raw(required=True, description='The symptoms that are shown',
                              example={"Fever": 1, "Cough": 0, "Diarrhoea": -1}),
    'likelihoods': fields.Raw(required=True, description='The likelihoods to be diagnosed',
                              example={"Rabies": {"Fever": 0.6, "Cough": 0.1, "Diarrhoea": 0.1},
                                       "Cold": {"Fever": 0.9, "Cough": 0.9, "Diarrhoea": 0.1}}),
    'priors': fields.Raw(required=False, description='The priors to be diagnosed', example={"Rabies": 20, "Cold": 80}),
    'animal': fields.String(required=False, description='The animal to be diagnosed', example='Dog')})


class getHelper:
    """
    A class used to get data from the Excel workbook
    """

    def __init__(self, *args, **kwargs):
        super(getHelper, self).__init__(*args, **kwargs)

    # A function used to collect the data from the Excel workbook which I manually formatted to ensure the data works.
    def get_disease_wiki_ids(self, animal):
        """
        A function used to get the WikiData IDs for the diseases
        :param animal:
        :return:
        """
        # Empty dictionary to store the WikiData IDs
        wiki_ids = {}
        # Get the list of diseases
        for disease in self.get_diseases(animal):
            # Loop through the diseases in the Excel sheet
            for row in wb['Disease_Codes'].rows:
                # If the disease is found, add it to the dictionary
                if row[0].value == disease:
                    wiki_ids[disease] = row[1].value
        # Return the dictionary
        return wiki_ids

    # A function used to collect the data from the Excel workbook which I manually formatted to ensure the data works.
    def get_likelihood_data(self, animal):
        """
        A function used to get the likelihood data from the Excel workbook
        :param animal:
        :return:
        """
        # Load the correct Excel sheet
        ws = wb[animal]
        # Get the list of signs and diseases
        signs = self.get_signs(animal)
        diseases = self.get_diseases(animal)
        # Dictionary which stores the likelihoods of each disease
        likelihoods = {}
        # Loop through all rows in the workbook
        for i, row in enumerate(ws.rows):
            # Skip the first row as it is just the headers
            if i == 0:
                continue
            # Dictionary which stores the likelihoods of each sign for the current disease
            current_disease_likelihoods = {}
            # Loop through all cells in each row except the first one as it is the disease name
            for j, cell in enumerate(row[1:]):
                chance = cell.value
                current_disease_likelihoods[signs[j]] = chance
            likelihoods[diseases[i - 1]] = current_disease_likelihoods
        return likelihoods

    @staticmethod
    def get_diseases(animal):
        """
        A function used to get the diseases from the Excel workbook
        :param animal:
        :return:
        """
        # List which stores every given disease
        diseases = []
        ws = wb[animal]
        # Populate the list of diseases
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
            for cell in row:
                diseases.append(cell.value)
        return diseases

    @staticmethod
    def get_signs(animal):
        """
        A function used to get the signs from the Excel workbook
        :param animal:
        :return:
        """
        # List which stores every given sign
        signs = []
        ws_signs = wb[animal]
        # Populate the list of signs
        for col in ws_signs.iter_cols(min_col=2, max_row=1):
            for cell in col:
                signs.append(cell.value)
        return signs

    @staticmethod
    def get_sign_names_and_codes(animal):
        """
        Get the full name and Wikidata code for each sign
        :param animal:
        :return:
        """
        ws = wb[animal + '_Abbr']
        full_sign_data = {}
        medical_name = {}
        wikidata_code = {}
        for row in ws.rows:
            medical_name["name"] = row[1].value
            wikidata_code["code"] = row[2].value
            full_sign_data[row[0].value] = {**medical_name, **wikidata_code}
        return full_sign_data


class diagnosisHelper:
    """
    A class used to validate the data provided by the user and perform the Bayes Theorem calculations to determine the
    probability of each disease
    """

    @staticmethod
    def validate_priors(priors, diseases):
        """
        A function used to validate the priors provided by the user
        :param priors:
        :param diseases:
        :return:
        """
        provided_keys = []
        for key in priors.keys():
            if key not in diseases:
                raise BadRequest(f"Disease '{key}' is not a valid disease. Please use a valid disease from {diseases}.")
            provided_keys.append(key)

        for disease in diseases:
            if disease not in provided_keys:
                raise BadRequest(
                    f"Missing '{disease}' in priors. Please provide a prior likelihood value for all diseases.")

        total_value = sum(priors.values())
        if total_value != 100:
            raise BadRequest(f"Priors must add up to 100. Currently they add up to {total_value}.")

        return priors

    @staticmethod
    def validate_likelihoods(likelihoods, diseases, signs):
        """
        A function used to validate the likelihoods provided by the user
        :param likelihoods:
        :param diseases:
        :param signs:
        :return:
        """
        provided_keys = []
        likelihoods = likelihoods
        for key in likelihoods.keys():
            if key not in diseases:
                raise BadRequest(f"Disease '{key}' in \'likelihoods\' is not a valid disease.")
            provided_keys.append(key)

        for disease in diseases:
            if disease not in provided_keys:
                raise BadRequest(
                    f"Missing '{disease}' in likelihoods. Please provide a likelihood value for all diseases.")

        for disease in likelihoods:
            current_likelihoods = likelihoods[disease]
            provided_keys = []
            for key in current_likelihoods.keys():
                if key not in signs:
                    raise BadRequest(f"Sign '{key}' in {disease} within \'likelihoods\' is not a valid sign. "
                                     f"Please use a valid signs from {signs}.")
                provided_keys.append(key)

            for sign in signs:
                if sign not in provided_keys:
                    raise BadRequest(
                        f"Missing '{sign}' in likelihoods for disease '{disease}'. Please provide a likelihood value"
                        f"for all signs.")

            for sign in current_likelihoods:
                if (not current_likelihoods[sign] > 0) or (current_likelihoods[sign] >= 1):
                    raise BadRequest(f"Likelihood for sign '{sign}' in disease '{disease}' is not a valid value. "
                                     f"Please use a value greater than 0 and less than 1.")

        return likelihoods

    @staticmethod
    def calculate_results(diseases, likelihoods, shown_signs, priors):
        """
        A function used to calculate the results of the Bayes Theorem
        :param diseases:
        :param likelihoods:
        :param shown_signs:
        :param priors:
        :return:
        """
        results = {}
        for disease in diseases:
            chain_probability = 1.0
            current_likelihoods = likelihoods[disease]
            for s in shown_signs:
                presence = shown_signs[s]
                if presence == 1:
                    chain_probability *= current_likelihoods[s]
                elif presence == -1:
                    chain_probability *= (1 - current_likelihoods[s])
            posterior = chain_probability * priors[disease]
            results[disease] = posterior * 100
        return results

    @staticmethod
    def normalise(results):
        """
        A function used to normalise the results of the Bayes Theorem calculations
        :param results:
        :return:
        """
        normalised_results = {}
        summed_results = sum(results.values())
        for r in results:
            value = results[r]
            norm = value / summed_results
            normalised_results[r] = norm * 100
        return normalised_results

    @staticmethod
    def get_default_priors(diseases):
        """
        A function used to generate equal priors if the user does not provide any
        :param diseases:
        :return priors:
        """
        priors = {}
        for disease in diseases:
            priors[disease] = 100 / len(diseases)
        return priors


@api.route('/api/diagnose/', methods=['POST'])
@api.doc(responses={200: 'OK', 400: 'Bad Request', 500: 'Internal Server Error'},
         description='<h1>Description</h1><p>This endpoint takes a <a href="https://developer.mozilla.org/'
                     'en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> object containing the species of animal and '
                     'a list of signs,and optionally a list of prior likelihood values. It then returns a list '
                     'of diseases and their likelihood of being the cause of the signs.</p> \n \n'
                     '<p>The <a href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a'
                     '> object must contain both "animal" and "signs" and can optionally contain '
                     '"priors" and "likelihoods"</p> \n \n<h1> Parameters</h1><p>animal:  You can use the'
                     '/api/data/valid_animals GET method to find out which animals are available for '
                     'diagnosis.</p>\n \n<p>signs: \'All signs detailed in the GET method '
                     '/api/data/full_sign_data/\'animal\' must be included. The data must be formatted as a '
                     '<a href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> '
                     'object,the key must be a string and the value of each sign must be 1 0, or -1. '
                     '1 means the sign is '
                     'present, 0 means the sign is not observed, but may still be present, and -1 means the sign is '
                     'not present.</p> \n \n<p>priors: This is an optional parameter which does not need to be passed '
                     'in the payload.It is used to alter the prior likelihoods of diseases occurring which influences '
                     'the outcome of the Bayes algorithm. If this is not included, the algorithm assumes equal prior '
                     'likelihoods. The list of diseases and their prior likelihoods must add up to 100.Data for the '
                     'required parameters can be returned by the GET Method at /api/data/full_disease_data/\'animal\' '
                     'which returns each disease as the key and the corresponding <a href="https://www.wikidata.org/">'
                     'WikiData ID</a> as the value. </p>\n \nlikelihoods: This is an optional parameter which does not '
                     'need to be passed in the payload. It is used to alter the likelihoods of signs being present for '
                     'each disease. If this is not included, the algorithm assumes equal likelihoods. The list of '
                     'diseases and their likelihoods must only contain values between 0 and 1, non-inclusive. Data '
                     'for the required parameters can be returned by the GET Method at /api/data/matrix/\'animal\' '
                     'which returns the default matrix the Bayesian algorithm uses. Alternatively the required signs '
                     'and diseases can be obtained via the /api/data/full_animal_data/\'animal\' endpoint.</p> \n \n'
                     '<p>WARNING: Providing your own likelihoods and priors can result in the algorithm returning '
                     'incorrect results. Use these features with caution if the values you provide are not based on '
                     'research. </p> <p>NOTE: likelihood values in the example payload are randomly generated '
                     'and as such should not be used to provide accurate results\n \n <h1> Example <a '
                     'href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> '
                     'object:</h1> \n \n  {\n"animal": "Cattle", \n\"signs\": {"Anae": 0, \"Anrx\": 1, \"Atax\": 0, '
                     '\"Const\": 0, \"Diarr\": 0, \"Dysnt\": 1, \"Dyspn\": 0, \"Icter\": 0, \"Lymph\": -1,\"Pyrx\": 0, '
                     '\"Stare\": 0, \"Stunt\": 0, \"SV_Oedm\": 1, \"Weak\": 0, \"Wght_L\": 0},\n '
                     '\"priors\": { \"Anthrax\": 5, \"Babesiosis\": 5, \"Blackleg\": 5, \"CBPP\": 5, '
                     '\"Colibacillosis\": 5, \"Cowdriosis\": 5,\"FMD\": 5,\"Fasciolosis\": 5,\"LSD\": 5, '
                     '\"Lungworm\":25,\"Pasteurollosis\": 5,\"PGE / GIT parasite\": 5,\"Rabies\": 5, '
                     '\"Trypanosomosis\": 5,\"Tuberculosis\": 5,\"ZZ_Other\": 5},\n'
                     '"likelihoods": { '
                     '    "Anthrax": {'
                     '      "Anae": 0.2668,'
                     '      "Anrx": 0.4023,'
                     '      "Atax": 0.0085,'
                     '      "Const": 0.185,'
                     '      "Diarr": 0.7954,'
                     '      "Dysnt": 0.0888,'
                     '      "Dyspn": 0.7864,'
                     '      "Icter": 0.2016,'
                     '      "Lymph": 0.6453,'
                     '      "Pyrx": 0.8447,'
                     '      "SV_Oedm": 0.0814,'
                     '      "Stare": 0.5038,'
                     '      "Stunt": 0.7275,'
                     '      "Weak": 0.0771,'
                     '      "Wght_L": 0.6383'
                     '    },'
                     '    "Babesiosis": {'
                     '      "Anae": 0.9315,'
                     '      "Anrx": 0.1084,'
                     '      "Atax": 0.5898,'
                     '      "Const": 0.7412,'
                     '      "Diarr": 0.2778,'
                     '      "Dysnt": 0.5351,'
                     '      "Dyspn": 0.2371,'
                     '      "Icter": 0.1604,'
                     '      "Lymph": 0.5144,'
                     '      "Pyrx": 0.8542,'
                     '      "SV_Oedm": 0.8295,'
                     '      "Stare": 0.8088,'
                     '      "Stunt": 0.9521,'
                     '      "Weak": 0.1691,'
                     '      "Wght_L": 0.6912'
                     '    },'
                     '    "Blackleg": {'
                     '      "Anae": 0.2589,'
                     '      "Anrx": 0.1933,'
                     '      "Atax": 0.503,'
                     '      "Const": 0.3529,'
                     '      "Diarr": 0.5158,'
                     '      "Dysnt": 0.9027,'
                     '      "Dyspn": 0.3662,'
                     '      "Icter": 0.6998,'
                     '      "Lymph": 0.3139,'
                     '      "Pyrx": 0.2509,'
                     '      "SV_Oedm": 0.735,'
                     '      "Stare": 0.1036,'
                     '      "Stunt": 0.5706,'
                     '      "Weak": 0.2188,'
                     '      "Wght_L": 0.0016'
                     '    },'
                     '    "CBPP": {'
                     '      "Anae": 0.4047,'
                     '      "Anrx": 0.9005,'
                     '      "Atax": 0.7564,'
                     '      "Const": 0.9049,'
                     '      "Diarr": 0.0014,'
                     '      "Dysnt": 0.0739,'
                     '      "Dyspn": 0.9572,'
                     '      "Icter": 0.2247,'
                     '      "Lymph": 0.0228,'
                     '      "Pyrx": 0.8588,'
                     '      "SV_Oedm": 0.2807,'
                     '      "Stare": 0.6536,'
                     '      "Stunt": 0.1723,'
                     '      "Weak": 0.0436,'
                     '      "Wght_L": 0.7293'
                     '    },'
                     '    "Colibacillosis": {'
                     '      "Anae": 0.5151,'
                     '      "Anrx": 0.6292,'
                     '      "Atax": 0.1373,'
                     '      "Const": 0.4801,'
                     '      "Diarr": 0.5341,'
                     '      "Dysnt": 0.8744,'
                     '      "Dyspn": 0.5039,'
                     '      "Icter": 0.7543,'
                     '      "Lymph": 0.4948,'
                     '      "Pyrx": 0.404,'
                     '      "SV_Oedm": 0.5243,'
                     '      "Stare": 0.1538,'
                     '      "Stunt": 0.6301,'
                     '      "Weak": 0.8463,'
                     '      "Wght_L": 0.1069'
                     '    },'
                     '    "Cowdriosis": {'
                     '      "Anae": 0.1476,'
                     '      "Anrx": 0.7902,'
                     '      "Atax": 0.7319,'
                     '      "Const": 0.7912,'
                     '      "Diarr": 0.7395,'
                     '      "Dysnt": 0.6757,'
                     '      "Dyspn": 0.1096,'
                     '      "Icter": 0.3056,'
                     '      "Lymph": 0.185,'
                     '      "Pyrx": 0.3019,'
                     '      "SV_Oedm": 0.4824,'
                     '      "Stare": 0.0284,'
                     '      "Stunt": 0.5792,'
                     '      "Weak": 0.1685,'
                     '      "Wght_L": 0.2448'
                     '    },'
                     '    "FMD": {'
                     '      "Anae": 0.2977,'
                     '      "Anrx": 0.871,'
                     '      "Atax": 0.963,'
                     '      "Const": 0.1151,'
                     '      "Diarr": 0.9518,'
                     '      "Dysnt": 0.6154,'
                     '      "Dyspn": 0.6984,'
                     '      "Icter": 0.4119,'
                     '      "Lymph": 0.3372,'
                     '      "Pyrx": 0.6846,'
                     '      "SV_Oedm": 0.0385,'
                     '      "Stare": 0.6973,'
                     '      "Stunt": 0.4824,'
                     '      "Weak": 0.6523,'
                     '      "Wght_L": 0.614'
                     '    },'
                     '    "Fasciolosis": {'
                     '      "Anae": 0.0095,'
                     '      "Anrx": 0.977,'
                     '      "Atax": 0.1164,'
                     '      "Const": 0.1311,'
                     '      "Diarr": 0.0569,'
                     '      "Dysnt": 0.1604,'
                     '      "Dyspn": 0.0823,'
                     '      "Icter": 0.4349,'
                     '      "Lymph": 0.6208,'
                     '      "Pyrx": 0.9651,'
                     '      "SV_Oedm": 0.1108,'
                     '      "Stare": 0.8521,'
                     '      "Stunt": 0.5637,'
                     '      "Weak": 0.1865,'
                     '      "Wght_L": 0.7242'
                     '    },'
                     '    "LSD": {'
                     '      "Anae": 0.0883,'
                     '      "Anrx": 0.3197,'
                     '      "Atax": 0.9291,'
                     '      "Const": 0.6622,'
                     '      "Diarr": 0.631,'
                     '      "Dysnt": 0.6657,'
                     '      "Dyspn": 0.974,'
                     '      "Icter": 0.153,'
                     '      "Lymph": 0.8023,'
                     '      "Pyrx": 0.2831,'
                     '      "SV_Oedm": 0.4065,'
                     '      "Stare": 0.2959,'
                     '      "Stunt": 0.3347,'
                     '      "Weak": 0.5577,'
                     '      "Wght_L": 0.9966'
                     '    },'
                     '    "Lungworm": {'
                     '      "Anae": 0.4765,'
                     '      "Anrx": 0.7152,'
                     '      "Atax": 0.6797,'
                     '      "Const": 0.938,'
                     '      "Diarr": 0.6535,'
                     '      "Dysnt": 0.9752,'
                     '      "Dyspn": 0.6245,'
                     '      "Icter": 0.4954,'
                     '      "Lymph": 0.4811,'
                     '      "Pyrx": 0.3523,'
                     '      "SV_Oedm": 0.2446,'
                     '      "Stare": 0.6479,'
                     '      "Stunt": 0.7919,'
                     '      "Weak": 0.7959,'
                     '      "Wght_L": 0.0079'
                     '    },'
                     '    "PGE / GIT parasite": {'
                     '      "Anae": 0.2108,'
                     '      "Anrx": 0.078,'
                     '      "Atax": 0.6913,'
                     '      "Const": 0.664,'
                     '      "Diarr": 0.5129,'
                     '      "Dysnt": 0.1369,'
                     '      "Dyspn": 0.27,'
                     '      "Icter": 0.9672,'
                     '      "Lymph": 0.8923,'
                     '      "Pyrx": 0.1874,'
                     '      "SV_Oedm": 0.7678,'
                     '      "Stare": 0.4827,'
                     '      "Stunt": 0.6827,'
                     '      "Weak": 0.8251,'
                     '      "Wght_L": 0.4358'
                     '    },'
                     '    "Pasteurollosis": {'
                     '      "Anae": 0.5699,'
                     '      "Anrx": 0.8898,'
                     '      "Atax": 0.7183,'
                     '      "Const": 0.9549,'
                     '      "Diarr": 0.4316,'
                     '      "Dysnt": 0.5618,'
                     '      "Dyspn": 0.4728,'
                     '      "Icter": 0.5825,'
                     '      "Lymph": 0.2789,'
                     '      "Pyrx": 0.6803,'
                     '      "SV_Oedm": 0.1556,'
                     '      "Stare": 0.1606,'
                     '      "Stunt": 0.5968,'
                     '      "Weak": 0.4604,'
                     '      "Wght_L": 0.1261'
                     '    },'
                     '    "Rabies": {'
                     '      "Anae": 0.021,'
                     '      "Anrx": 0.1521,'
                     '      "Atax": 0.99,'
                     '      "Const": 0.553,'
                     '      "Diarr": 0.6056,'
                     '      "Dysnt": 0.123,'
                     '      "Dyspn": 0.5579,'
                     '      "Icter": 0.8676,'
                     '      "Lymph": 0.4572,'
                     '      "Pyrx": 0.4503,'
                     '      "SV_Oedm": 0.8823,'
                     '      "Stare": 0.6489,'
                     '      "Stunt": 0.0886,'
                     '      "Weak": 0.5222,'
                     '      "Wght_L": 0.4999'
                     '    },'
                     '    "Trypanosomosis": {'
                     '      "Anae": 0.3602,'
                     '      "Anrx": 0.6961,'
                     '      "Atax": 0.898,'
                     '      "Const": 0.6043,'
                     '      "Diarr": 0.5259,'
                     '      "Dysnt": 0.3489,'
                     '      "Dyspn": 0.6187,'
                     '      "Icter": 0.3156,'
                     '      "Lymph": 0.0677,'
                     '      "Pyrx": 0.0405,'
                     '      "SV_Oedm": 0.0257,'
                     '      "Stare": 0.5779,'
                     '      "Stunt": 0.9064,'
                     '      "Weak": 0.0779,'
                     '      "Wght_L": 0.9971'
                     '    },'
                     '    "Tuberculosis": {'
                     '      "Anae": 0.2207,'
                     '      "Anrx": 0.7513,'
                     '      "Atax": 0.9402,'
                     '      "Const": 0.5418,'
                     '      "Diarr": 0.9419,'
                     '      "Dysnt": 0.0402,'
                     '      "Dyspn": 0.6023,'
                     '      "Icter": 0.399,'
                     '      "Lymph": 0.6675,'
                     '      "Pyrx": 0.5587,'
                     '      "SV_Oedm": 0.4947,'
                     '      "Stare": 0.59,'
                     '      "Stunt": 0.1479,'
                     '      "Weak": 0.8633,'
                     '      "Wght_L": 0.085'
                     '    },'
                     '    "ZZ_Other": {'
                     '      "Anae": 0.6361,'
                     '      "Anrx": 0.1729,'
                     '      "Atax": 0.3008,'
                     '      "Const": 0.6395,'
                     '      "Diarr": 0.3737,'
                     '      "Dysnt": 0.858,'
                     '      "Dyspn": 0.9196,'
                     '      "Icter": 0.3255,'
                     '      "Lymph": 0.4681,'
                     '      "Pyrx": 0.7769,'
                     '      "SV_Oedm": 0.0234,'
                     '      "Stare": 0.9337,'
                     '      "Stunt": 0.3675,'
                     '      "Weak": 0.1286,'
                     '      "Wght_L": 0.0943'
                     '    }'
                     ' }\n \t }\n } ')
class diagnose(Resource):
    """
    This class is used to handle the diagnosis of the animal
    """

    def __init__(self, *args, **kwargs):
        """
        This method is used to initialize the class and set the helper classes to be used throughout the diagnosis process
        :param args:
        :param kwargs:
        """
        self.gh = getHelper()
        self.dh = diagnosisHelper()
        super(diagnose, self).__init__(*args, **kwargs)

    @staticmethod
    def options() -> Response:
        # This is required to allow the OPTIONS method to be used for CORS
        return jsonify({'status': 'ok'})

    @api.expect(diagnosis_payload_model, validate=True)
    def post(self) -> Response:

        data: [str, Union[str, Dict[str, int], None]] = request.get_json()
        animal: str = data['animal']

        animal = validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /api/data/valid_animals.', 'status': 404}, 404

        valid_signs: List[str] = self.gh.get_signs(animal)
        diseases: List[str] = self.gh.get_diseases(animal)
        wiki_ids: Dict[str, str] = self.gh.get_disease_wiki_ids(animal)

        # Check if the likelihoods are included in the API request data
        if data.get('likelihoods') is not None:
            likelihoods = self.dh.validate_likelihoods(data['likelihoods'], diseases, valid_signs)
        else:
            likelihoods: Dict[str, Dict[str, float]] = self.gh.get_likelihood_data(animal)

        # Get the signs from the API request data
        shown_signs: Dict[str, int] = data['signs']
        if set(shown_signs.keys()) != set(valid_signs):
            raise BadRequest(f'Invalid signs: {list(set(shown_signs.keys()) - set(valid_signs))}. '
                             f'Please use valid sign from /api/data/valid_signs/{animal}.')

        valid_sign_values = [0, 1, -1]
        for x, value in enumerate(shown_signs.values()):
            if value not in valid_sign_values:
                sign = list(shown_signs.keys())[x]
                raise BadRequest(f'Error with value of {sign}: {value}. Sign values must be either -1, 0 or 1')

        # Check if the priors are included in the API request data
        if data.get('priors') is not None:
            priors = self.dh.validate_priors(data.get('priors'), diseases)
        else:
            priors = self.dh.get_default_priors(diseases)

        # Perform calculations and normalisation
        results: Dict[str, float] = self.dh.calculate_results(diseases, likelihoods, shown_signs, priors)
        normalised_results: Dict[str, float] = self.dh.normalise(results)

        return jsonify({'results': normalised_results, 'wiki_ids': wiki_ids})


@api.route('/api/custom_diagnose', methods=['POST'])
@api.doc(responses={200: 'OK', 400: 'Bad Request', 500: 'Internal Server Error'},
         description='<h1>Description</h1><p>This endpoint takes a <a '
                     'href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> '
                     'object containing the list of possible diseases for the animal you wish to diagnose,'
                     'the list of signs you intend to input,'
                     'the signs paired with the the value relating to the presence of the sign '
                     'and optionally a list of prior likelihood values for the disease to occur in the animal and the '
                     'type of animal it is. It then returns a list of diseases and their likelihood of being the cause '
                     'of the signs.</p> \n \n<p>The '
                     '<a href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> object '
                     'must contain both "diseases", "signs" and "shown_signs" and can optionally contain "priors" and '
                     '"likelihoods"</p> \n \n<h1>'
                     'Parameters</h1><p>animal:  You can input any string, this is currently arbitrary and unused, '
                     'but may in future be used to store data to gather information.</p>\n \n<p>signs: This is a list '
                     'of signs that will be used to diagnose the animal. The list of signs must be the same as the '
                     'list of shown_signs for the animal you are diagnosing. </p>\n \n<p>shown_signs: This is a list '
                     'of signs that are present in the animal. The list of signs must be the same as the list of '
                     'signs for the animal you are diagnosing. The value of each sign must be either -1, 0 or 1. '
                     '-1 means the sign is not present, 0 means the sign is not observed  and 1 means the '
                     'sign is present in the animal.</p> \n \n'
                     '<p>likelihoods: This is the data which will be provided to the algorithm to calculate the '
                     'likelihood of each disease being the cause of the signs. The likelihoods must be formatted '
                     'as a dictionary with the disease as the key and a dictionary of signs and their likelihoods '
                     'as the value. The signs must be formatted as a dictionary with the sign as the key and the '
                     'likelihood of the sign being present for the disease as the value. The likelihoods must '
                     'be a float between 0 and 1. </p>\n \n'
                     '<p>priors: This is an optional parameter which does not need to be passed in '
                     'the payload.It is used to alter the prior likelihoods of diseases occurring which influences the '
                     'outcome of the Bayes algorithm. If this is not included, the algorithm assumes equal prior '
                     'likelihoods. The list of diseases and their prior likelihoods must add up to 100.'
                     'The priors must be formatted as a dictionary with the disease as the key and the prior '
                     'likelihood of the disease as the value. The prior likelihoods must be a float between 0 and 100'
                     '. You must include every disease in the list of diseases for the animal you are diagnosing. '
                     ' </p>\n \n <h1> '
                     'Example <a href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON'
                     '</a> '
                     'object:</h1> \n \n  {\n \"diseases\": [\"Rabies\",\"Cold\"], \n'
                     '\"signs\": [\"Fever\", \"Cough\", \"Diarrhoea\"],\n '
                     '\"shown_signs\": {\"Fever\": 1, \"Cough\": 0 \"Diarrhoea\": -1},\n '
                     '"likelihoods": {"Rabies": {"Fever": 0.6, "Cough": 0.1, "Diarrhoea": 0.1},"Cold": '
                     '{"Fever": 0.9,"Cough": 0.9,"Diarrhoea": 0.1}},\n '
                     '"priors": {"Rabies": 20,"Cold": 80 },\n "animal": "Dog" \n}')
class custom_diagnose(Resource):
    """
    This class is used to create the custom_diagnose endpoint.
    """

    def __init__(self, *args, **kwargs):
        # Initialise the helper classes
        self.gh = getHelper()
        self.dh = diagnosisHelper()
        super(custom_diagnose, self).__init__(*args, **kwargs)

    @staticmethod
    def options() -> Response:
        # This is required to allow the OPTIONS method to be used for CORS
        return jsonify({'status': 'ok'})

    @api.expect(custom_diagnosis_payload_model, validate=True)
    def post(self):
        # This is the POST method for the custom_diagnose endpoint which allows the user to input their own data,
        # meaning the API can be used in a larger number of contexts.

        data: [str, Union[str, Dict[str, int], None]] = request.get_json()
        shown_signs: Dict[str, int] = data.get('shown_signs')
        likelihoods: Dict[str, Dict[str, float]] = data.get('likelihoods')
        diseases: List[str] = data.get('diseases')
        priors: Dict[str, float] = data.get('priors')
        sign_list: List[str] = data.get('signs')

        valid_sign_values = [0, 1, -1]
        # Check if the signs values are all valid
        for x, value in enumerate(shown_signs.values()):
            if value not in valid_sign_values:
                sign = list(shown_signs.keys())[x]
                raise BadRequest(f'Error with value of {sign}: {value}. Sign values must be either -1, 0 or 1')

        # Check to make sure the likelihoods are valid
        self.dh.validate_likelihoods(likelihoods, diseases, sign_list)

        # Check to make sure the priors are valid
        if priors is not None:
            self.dh.validate_priors(priors, diseases)
        else:
            priors = self.dh.get_default_priors(diseases)

        results: Dict[str, float] = self.dh.calculate_results(diseases, likelihoods, shown_signs, priors)
        normalised_results: Dict[str, float] = self.dh.normalise(results)

        return jsonify({'results': normalised_results})


@api.route('/api/data/full_animal_data/<string:animal>')
@api.doc(example='Goat', required=True,
         responses={200: 'OK', 400: 'Bad Request', 404: 'Not Found', 500: 'Internal Server Error'},
         params={'animal': 'The species of animal you wish to retrieve signs and diseases for. This must be a valid '
                           'animal as returned by /api/data/valid_animals. \n \n'},
         description='<h1>Description</h1><p>This endpoint returns a '
                     '<a href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> '
                     'object containing diagnosable diseases with their corresponding '
                     '<a href="https://www.wikidata.org/">WikiData IDs</a> as well as the valid signs associated with '
                     'the animal, their full medical terminology in English, and their corresponding '
                     '<a href="https://www.wikidata.org/">WikiData IDs</a> (if they exist).</p>'
                     '<h1>URL Parameters</h1><ul><li><p>animal: The species of animal you wish to retrieve signs and '
                     'diseases for. This must be a valid animal as returned by /api/data/valid_animals. '
                     '</p></li></ul>\n \n ')
class getRequiredInputData(Resource):
    """
    This class is used to create the full_animal_data endpoint which returns the required input data for the
    diagnose endpoint.
    """

    def __init__(self, *args, **kwargs):
        """
        This function is used to initialise the getRequiredInputData class and instantiate the helper class.
        :param args:
        :param kwargs:
        """
        self.gh = getHelper()
        super(getRequiredInputData, self).__init__(*args, **kwargs)

    def get(self, animal):
        # This is the GET method for the full_animal_data endpoint which returns the required input data for the
        # diagnose endpoint.

        animal = validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /api/data/valid_animals.', 'status': 404}, 404

        return jsonify(
            {'diseases': self.gh.get_disease_wiki_ids(animal), 'signs': self.gh.get_sign_names_and_codes(animal)})


@api.hide
@api.route('/api/data/matrix/<string:animal>')
@api.doc(example='Goat', required=True,
         responses={200: 'OK', 400: 'Bad Request', 404: 'Not Found', 500: 'Internal Server Error'}, params={
        'animal': 'The species of animal you wish to retrieve the disease sign matrix for. This must be a valid '
                  'animal as returned by /api/data/valid_animals. \n \n'},
         description='<h1>Description</h1><p>This endpoint returns a '
                     '<a href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> '
                     'object containing the disease-sign Bayesian matrix for the given animal. This matrix contains '
                     'the likelihoods of each sign being present for each disease.</p><h1>URL Parameters</h1><ul><li>'
                     '<p>animal: The species of animal you wish to retrieve the disease-sign matrix for. '
                     'This must be a valid animal as returned by /api/data/valid_animals. </p></li></ul>\n \n ')
class getDiseaseSignMatrix(Resource):
    """
    This class is used to create the matrix endpoint which returns the disease sign matrix for the given animal. It is
    hidden from the Swagger UI as it is not intended to be used by the user, it is only to be used by the developer
    or specific users permitted in order to update the data on mobile apps which need to perform calculations offline.
    """

    def __init__(self, *args, **kwargs):
        self.gh = getHelper()
        super(getDiseaseSignMatrix, self).__init__(*args, **kwargs)

    def get(self, animal):
        # This is the GET method for the matrix endpoint which returns the disease sign matrix for the given animal.
        # It is hidden from the Swagger UI as it is not intended to be used by the user, it is only to be used by the
        # developer or specific users permitted in order to update the data on mobile apps which need to perform
        # calculations offline.

        animal = validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /api/data/valid_animals.', 'status': 404}, 404

        return jsonify(self.gh.get_likelihood_data(animal))


@api.route('/api/data/example_matrix/<string:animal>')
@api.doc(example='Sheep', required=True,
         responses={200: 'OK', 400: 'Bad Request', 404: 'Not Found', 500: 'Internal Server Error'}, params={
        'animal': 'The species of animal you wish to retrieve the disease sign matrix for. This must be a valid '
                  'animal as returned by /api/data/valid_animals. \n \n'},
         description='<h1>Description</h1><p>This endpoint returns a '
                     '<a href="https://developer.mozilla.org/en-US/docs/Learn/JavaScript/Objects/JSON">JSON</a> '
                     'object containing an example disease-sign Bayesian matrix for the given animal. '
                     'This matrix contains randomly generated '
                     'likelihoods of each sign being present for each disease.</p><h1>URL Parameters</h1><ul><li>'
                     '<p>animal: The species of animal you wish to retrieve the disease-sign matrix for. '
                     'This must be a valid animal as returned by /api/data/valid_animals. </p></li></ul>\n \n ')
class getExampleMatrix(Resource):
    """
    This class is used to create the example_matrix endpoint which returns a randomly generated disease sign matrix
    for the given animal. It is used for documentation and giving context to users of the API as the data used in
    the diagnosis process is not available to the public.
    """

    def __init__(self, *args, **kwargs):
        self.gh = getHelper()
        super(getExampleMatrix, self).__init__(*args, **kwargs)

    def get(self, animal):
        # This is the GET method for the example_matrix endpoint

        animal = validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /api/data/valid_animals.', 'status': 404}, 404

        data = self.gh.get_likelihood_data(animal)

        # Replace correct data with random data
        for disease in data:
            for sign in data[disease]:
                data[disease][sign] = round(random.uniform(0.00001, 0.99999), 4)

        return jsonify(data)


@api.route('/api/data/valid_animals')
@api.doc(responses={200: 'OK', 500: 'Internal Server Error'}, description='<h1>Description</h1>'
                                                                          '<p>This endpoint returns a <a '
                                                                          'href="https://developer.mozilla.org/en-US/docs'
                                                                          '/Learn/JavaScript'
                                                                          '/Objects/JSON">JSON</a> object containing the '
                                                                          'names of animal species which are available for '
                                                                          'diagnosis in the /api/diagnose POST method '
                                                                          'below. </p>\n')
class getAnimals(Resource):
    """
    This class is used to create the valid_animals endpoint which returns the names of animals which are available for
    diagnosis in the /api/diagnose POST method
    """

    @staticmethod
    def get():
        # This is the GET method for the valid_animals endpoint

        # Return the names by getting every worksheet with doesn't contain _Abbr or _Codes
        # I'm using list comprehension as it is cleaner than a for loop, but syntacitically slightly more advanced

        return jsonify([name for name in wb.sheetnames if "_Abbr" not in name and "_Codes" not in name])


@api.route('/api/data/full_sign_data/<string:animal>')
@api.doc(required=True, responses={200: 'OK', 400: 'Bad Request', 404: 'Not Found', 500: 'Internal Server Error'},
         description='<h1>Description</h1>'
                     '<p>This endpoint returns a <a '
                     'href="https://developer.mozilla'
                     '.org/en-US/docs/Learn/JavaScript'
                     '/Objects/JSON">JSON</a> object '
                     'which contains the full medical '
                     'terminology for each sign'
                     ' in English as '
                     'well as the corresponding <a '
                     'href="https://www.wikidata.org'
                     '/">WikiData IDs</a> for the '
                     'signs ('
                     'if they exist).</p>'
                     '<h1>URL Parameters</h1>'
                     '<ul>'
                     '<li><p>animal: The species of '
                     'animal you wish to retrieve '
                     'signs and diseases for. This '
                     'must be '
                     'a valid '
                     'animal as returned by '
                     '/api/data/valid_animals.</p></li>'
                     '</ul>',
         params={'animal': 'The species of animal you wish to retrieve the data for. This must be a valid animal as '
                           'returned by /api/data/valid_animals. \n \n '})
class getSignCodesAndTerminology(Resource):
    """
    This class is used to create the full_sign_data endpoint which returns the full medical terminology for each sign
    in English as well as the corresponding WikiData IDs for the signs (if they exist).
    """

    def __init__(self, *args, **kwargs):
        self.gh = getHelper()
        super(getSignCodesAndTerminology, self).__init__(*args, **kwargs)

    def get(self, animal):
        # This is the GET method for the full_sign_data endpoint

        animal = validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /api/data/valid_animals.', 'status': 404}, 404

        return jsonify({'full_sign_data': self.gh.get_sign_names_and_codes(animal)})


@api.route('/api/data/full_disease_data/<string:animal>')
@api.doc(required=True, responses={200: 'OK', 400: 'Bad Request', 404: 'Not Found', 500: 'Internal Server Error'},
         description='<h1>Description</h1>'
                     '<p>This endpoint returns a <a '
                     'href="https://developer.mozilla'
                     '.org/en-US/docs/Learn/JavaScript'
                     '/Objects/JSON">JSON</a> object '
                     'which contains the possible '
                     'diseases for the given animal as '
                     'well as the corresponding <a '
                     'href="https://www.wikidata.org'
                     '/">WikiData IDs</a> (if they '
                     'exist).</p>'
                     '<h1>URL Parameters</h1>'
                     '<ul>'
                     '<li><p>animal: The species of '
                     'animal you wish to retrieve '
                     'signs and diseases for. This '
                     'must be '
                     'a valid '
                     'animal as returned by '
                     '/api/data/valid_animals.</p></li>'
                     '</ul>',
         params={'animal': 'The species of animal you wish to retrieve the data for. This must be a valid animal as '
                           'returned by'
                           '/api/data/valid_animals. \n \n'})
class getDiseaseCodes(Resource):
    """
    This class is used to create the full_disease_data endpoint which returns the possible diseases for the given
    animal as well as the corresponding WikiData IDs (if they exist).
    """

    def __init__(self, *args, **kwargs):
        self.gh = getHelper()
        super(getDiseaseCodes, self).__init__(*args, **kwargs)

    def get(self, animal):
        # This is the GET method for the full_disease_data endpoint

        animal = validate_animal(animal)
        if animal is False:
            return {'error': 'Invalid animal. Please use a valid animal '
                             'from /api/data/valid_animals.', 'status': 404}, 404

        return jsonify({'disease_codes': self.gh.get_disease_wiki_ids(animal)})


def validate_animal(animal):
    """
    This function is used to validate the animal parameter in the endpoints, it is a global function as it is used in
    multiple endpoints and in multiple functions in both the getHelper and diseaseHelper classes.
    :param animal:
    :return bool:
    """
    animal = animal.capitalize()
    if animal not in getAnimals().get().get_json():
        return False
    else:
        return animal


if __name__ == '__main__':
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
    app.debug = True
    app.run(port=5000)
