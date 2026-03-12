"""
One-time script to convert data.xlsx into data.json.
Re-run this whenever the Excel data changes.
"""

import json
import os
import sys

from openpyxl import load_workbook

wb = load_workbook(filename=os.path.join(sys.path[0], "data.xlsx"), read_only=True, data_only=True)

animals = [name for name in wb.sheetnames if "_Abbr" not in name and "_Codes" not in name]

# Build disease codes lookup from the shared sheet
all_disease_codes = {}
for row in wb['Disease_Codes'].rows:
    all_disease_codes[row[0].value] = row[1].value

data = {"animals": {}}

for animal in animals:
    ws = wb[animal]

    # Extract signs (header row, skip first column)
    signs = []
    diseases = []
    likelihoods = {}

    for i, row in enumerate(ws.rows):
        if i == 0:
            signs = [cell.value for cell in row[1:]]
            continue
        disease_name = row[0].value
        diseases.append(disease_name)
        likelihoods[disease_name] = {
            signs[j]: cell.value for j, cell in enumerate(row[1:])
        }

    # Disease wiki IDs for this animal
    disease_wiki_ids = {}
    for disease in diseases:
        if disease in all_disease_codes:
            disease_wiki_ids[disease] = all_disease_codes[disease]

    # Sign names and codes from _Abbr sheet
    ws_abbr = wb[animal + '_Abbr']
    sign_names_and_codes = {}
    for row in ws_abbr.rows:
        sign_names_and_codes[row[0].value] = {
            "name": row[1].value,
            "code": row[2].value
        }

    data["animals"][animal] = {
        "signs": signs,
        "diseases": diseases,
        "likelihoods": likelihoods,
        "disease_wiki_ids": disease_wiki_ids,
        "sign_names_and_codes": sign_names_and_codes
    }

wb.close()

output_path = os.path.join(sys.path[0], "data.json")
with open(output_path, "w") as f:
    json.dump(data, f, indent=2)

print(f"Wrote {output_path} with animals: {animals}")
